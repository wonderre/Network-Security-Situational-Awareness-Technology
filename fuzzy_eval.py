import  numpy as np
import pandas as pd
import csv
import openpyxl as op
import os
#计算权重向量
def weights_Count(array):
    m = len(array)                                    #获取指标个数
    n = len(array[0])
    R = np.linalg.matrix_rank(array)                #求判断矩阵的秩
    V, D = np.linalg.eig(array)                       #求判断矩阵的特征值和特征向量，V特征值，D特征向量；
    list1 = list(V)
    B = np.max(list1)#最大特征值
    index = list1.index(B)
    C = D[:, index]                            #对应特征向量
    #print("各向量权重向量Q为：")
    sum = np.sum(C)
    Q = C/sum                               #特征向量标准化
    #print(Q)                              #  输出权重向量
    return Q

# 获取评价数据
def get_DataFromExcel(count):
    f = csv.reader(open(r"./Mark/mark"+str(count)+".csv", 'r'))
    show = []
    for i in f:
        show.append(i)
    df=pd.DataFrame(show)
    #print(df)
    return df

# 模糊综合评价法(FCE)，输入准则权重、因素权重
def fuzzy_eval(criteria, eigen, count):
    # 量化评语（优秀、    良好、    一般、    较差、   非常差）
    score = [1, 0.8, 0.6, 0.4, 0.2]
    df = get_DataFromExcel(count)
    #print('单因素模糊综合评价：{}\n'.format(df))
    # 把单因素评价数据，拆解到5个准则中
    v1 = df.iloc[0:2, :].values
    v2 = df.iloc[2:7, :].values
    v3 = df.iloc[7:9, :].values

    vv = [v1, v2, v3]
    val = []
    num = len(eigen)
    #将字符串类型的数据转为float型才能进行计算
    for n in range(len(vv)):
        for i in range(vv[n].shape[0]):
            for j in range(vv[n].shape[1]):
                vv[n][i][j]=float(vv[n][i][j])
    for i in range(num):
        v = np.dot(eigen[i], vv[i])
        #print('准则{} , 矩阵积为：{}'.format(i + 1, v))
        val.append(v)

    # 目标层
    obj = np.dot(criteria, np.array(val))

    sum = obj.sum()
    for i in range(len(obj)):
        obj[i] =  obj[i]/sum

    print('目标层模糊综合评价：{}'.format(obj))
    # 综合评分
    eval = np.dot(np.array(obj), np.array(score).T)
    print('综合评价：{}'.format(eval))
    return obj, eval

def evaluation(obj):
    obj=list(obj)
    for i in range(len(obj)):
        obj[i]=obj[i].real
    max_data=max(obj)
    n=obj.index(max_data)
    eva=['优','良','中','差','危']
    return eva[n]

if __name__ == "__main__":

    first_Lev = np.array([[0.5, 0.3, 0.6],
                        [0.7, 0.5, 0.8],
                        [0.4, 0.2, 0.5]])
    first_Weig = weights_Count(first_Lev)

    second_Lev1 = np.array([[0.5, 0.3],
                            [0.7, 0.5]])
    second_Weig1 = weights_Count(second_Lev1)

    second_Lev2 = np.array([[0.5, 0.5, 0.4, 0.3, 0.4],
                [0.5, 0.5, 0.4, 0.3, 0.4],
                [0.6, 0.6, 0.5, 0.6, 0.3],
                [0.7, 0.7, 0.4, 0.5, 0.2],
                [0.6, 0.6, 0.7, 0.8, 0.5]])
    second_Weig2 = weights_Count(second_Lev2)

    second_Lev3 = np.array([[0.5, 0.3],
                            [0.7, 0.5]])
    second_Weig3 = weights_Count(second_Lev3)
    criteria = first_Weig
    eigen = []
    eigen.append(second_Weig1)
    eigen.append(second_Weig2)
    eigen.append(second_Weig3)
    df = pd.read_excel(r".\index\index.xlsx", "Sheet1")
    y = df.shape[0]
    count = 1

    df1 = pd.DataFrame(
            {
                "time":[],
                "eval":[]
            })
    df1.to_excel("predict.xlsx")
    time=df.loc[:,"Timestamp"]
    for i in range(y):
        obj, eval = fuzzy_eval(criteria, eigen, count)
        eva = evaluation(obj)
        timestamp=time[i]
        num_list = [timestamp, eval.real]
        bg = op.load_workbook("predict.xlsx")  # 应先将excel文件放入到工作目录下
        sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
        for i in range(1, len(num_list) + 1):
            sheet.cell(count+1, i + 1,
                       num_list[i - 1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第2行第2列
            bg.save("predict.xlsx")  # 对文件进行保存

        count += 1
        print('网络态势评估:'+eva+'\n')
data = pd.read_excel('predict.xlsx','Sheet1',index_col=0)
data.to_csv('predict.csv',encoding='utf-8')
os.remove('predict.xlsx')