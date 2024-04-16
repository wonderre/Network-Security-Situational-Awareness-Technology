import pandas as pd
import os
import openpyxl as op
import glob
#    数据包大小占比（PktAvg函数）：
#       你计算了数据集中数据包大小的占比情况，分别包括正常（Benign）和异常（非Benign）情况下小于等于32字节和大于100字节的数据包的比例。
#       通过统计数据包大小在两种情况下的比例，可以初步了解异常情况下是否有异常大小的数据包占比较高，从而可能表示攻击行为。
#    流的标准差长度（PktLenStd函数）：
#       你计算了流的标准差长度的平均值，分别针对正常和异常情况。
#       标准差长度的变化可能反映了流量的变化程度，异常情况下的标准差长度可能较大，表明流量的波动性增加。
#   流量持续时间（FlowDur函数）：
#       你计算了流量持续时间的平均值，同样分为正常和异常情况。
#       这可以帮助你了解流量持续时间的分布情况，异常情况下可能存在持续时间异常偏长或偏短的情况。
#   下载上传比率（Down_Up_Ratio函数）：
#      你计算了下载上传比率的平均值，同样分为正常和异常情况。
#       这可以帮助你了解流量的下载和上传行为之间的关系，异常情况下可能会有不同的下载上传比率分布。
#   恶意攻击的频率（Fre_mal函数）：
#       你计算了恶意攻击的频率，即异常情况在总流量中的比例。
#       这可以帮助你了解整个数据集中恶意攻击的程度，频率较高可能意味着网络受到了较多的攻击。
#   TCP和UDP数据包的占比（PktTu函数）：
#       你计算了TCP和UDP数据包的总体占比，以及在正常和异常情况下的占比情况。
#       这可以帮助你了解不同协议的数据包在整个流量中的分布情况，可能有助于发现异常协议使用情况。
# 数据包大小
def PktAvg(file_path):
    # 读取CSV文件为DataFrame
    df = pd.read_csv(file_path)
    # 从DataFrame中提取正常和异常数据
    B_df = df.loc[df['Label'] == 'Benign']
    # print(B_df)
    # 计算正常和异常数据的总数
    B_total = len(B_df)
    # print(B_total)
    N_df = df.loc[df['Label'] != 'Benign']
    N_total = len(N_df)
    # print(N_total)
    # 计算正常和异常数据中小于等于32字节和大于100字节的数据包数量
    B_total_Small = len(B_df.loc[df['Pkt Size Avg'] <= 32, ['Pkt Size Avg']])
    B_total_Big = len(B_df.loc[df['Pkt Size Avg'] > 100, ['Pkt Size Avg']])
    N_total_Small = len(N_df.loc[df['Pkt Size Avg'] <= 32, ['Pkt Size Avg']])
    N_total_Big = len(N_df.loc[df['Pkt Size Avg'] > 100, ['Pkt Size Avg']])
    T_total_Small=len(df.loc[df['Pkt Size Avg'] <= 32, ['Pkt Size Avg']])
    T_total_Big=len(B_df.loc[df['Pkt Size Avg'] >100, ['Pkt Size Avg']])
    if B_total!=0:
        B_ratio_Small = format(float(B_total_Small) / float(B_total), '.5f')
        B_ratio_Big = format(float(B_total_Big) / float(B_total), '.5f')
    else:
        B_ratio_Small=B_ratio_Big=0
    if N_total!=0:
        N_ratio_Small = format(float(N_total_Small) / float(N_total), '.5f')
        N_ratio_Big = format(float(N_total_Big) / float(N_total), '.5f')
    else:
        N_ratio_Small=N_ratio_Big=0
    T_ratio_Small=format(float(T_total_Small) / float(len(df)), '.5f')
    T_ratio_Big = format(float(T_total_Big) / float(len(df)), '.5f')
    return T_ratio_Small,T_ratio_Big,B_ratio_Small, B_ratio_Big, N_ratio_Small, N_ratio_Big


# 流的标准差长度
def PktLenStd(file_path):
    df = pd.read_csv(file_path)
    # 从DataFrame中提取正常和异常数据的标准差长度
    B_df = df.loc[df['Label'] == 'Benign', ['Pkt Len Std']]
    # 计算正常和异常数据的数量
    B_total = len(B_df)
    N_df = df.loc[df['Label'] != 'Benign', ['Pkt Len Std']]
    N_total = len(N_df)
    # 计算正常和异常数据的标准差长度的平均值
    if B_total!=0:
        B_PktLen = format(B_df["Pkt Len Std"].sum() / B_total, '.5f')
    else:
        B_PktLen=0
    if N_total!=0:
        N_PktLen = format(N_df["Pkt Len Std"].sum() / N_total, '.5f')
    else:
        N_PktLen=0
    # 计算总体的标准差长度的平均值
    T_PktLen=format(df["Pkt Len Std"].sum() / len(df), '.5f')
    return T_PktLen,B_PktLen, N_PktLen


# 流量持续时间
def FlowDur(file_path):
    df = pd.read_csv(file_path)
    B_df = df.loc[df['Label'] == 'Benign', ['Flow Duration']]
    B_total = len(B_df)
    N_df = df.loc[df['Label'] != 'Benign', ['Flow Duration']]
    N_total = len(N_df)
    # 计算正常和异常数据的流量持续时间的平均值
    if B_total!=0:
        B_Flowdur = format(B_df["Flow Duration"].sum() / B_total, '.5f')
    else:
        B_Flowdur=0
    if N_total!=0:
        N_Flowdur = format(N_df["Flow Duration"].sum() / N_total, '.5f')
    else:
        N_Flowdur=0
    # 计算总体的流量持续时间的平均值
    T_Flowdur=format(df["Flow Duration"].sum() / len(df), '.5f')
    return T_Flowdur,B_Flowdur, N_Flowdur

#下载上传比率
def Down_Up_Ratio(file_path):
    df = pd.read_csv(file_path)
    B_df = df.loc[df['Label'] == 'Benign', ['Down/Up Ratio']]
    B_total = len(B_df)
    N_df = df.loc[df['Label'] != 'Benign', ['Down/Up Ratio']]
    N_total = len(N_df)
    if B_total != 0:
        B_DownUp_Ratio = format(B_df["Down/Up Ratio"].sum() / B_total, '.5f')
    else:
        B_DownUp_Ratio = 0
    if N_total != 0:
        N_DownUp_Ratio = format(N_df["Down/Up Ratio"].sum() / N_total, '.5f')
    else:
        N_DownUp_Ratio = 0
    T_DownUp_Ratio=format(df["Down/Up Ratio"].sum() / len(df), '.5f')
    return T_DownUp_Ratio,B_DownUp_Ratio,N_DownUp_Ratio


# 恶意攻击的频率
def Fre_mal(file_path):
    df = pd.read_csv(file_path)
    # 计算总流量的数量
    total = len(df)
    # 从DataFrame中提取异常数据
    N_df = df.loc[df['Label'] != 'Benign']
    # 计算异常数据的数量
    N_total = len(N_df)
    # 计算异常数据在总流量中的频率
    N_freq = format(float(N_total) / float(total), '.5f')
    return N_freq


##计算TCP和UDP数据包的占比
def PktTu(file_path):
    label_type = []  # 攻击类型
    num = []  # 攻击数量
    frequency = []
    df = pd.read_csv(file_path)
    # 获取数据集中不同的攻击类型和它们出现的次数
    label = df.loc[:, "Label"].values
    for i in label:
        if i not in label_type:
            label_type.append(i)
            num.append(1)
        else:
            num[label_type.index(i)] += 1
    # 计算不同攻击类型的频率
    for n in range(len(label_type)):
        frequency.append(num[n] / len(label))
        #print(label_type[n] + "攻击频率：", frequency[n])
    # 计算TCP和UDP数据包的占比
    UDP_NUM = [0, 0]  # UDP_NUM[0]正常情况下UDP数据包的数量，UDP_NUM[1]是受到攻击的数量
    TCP_NUM = [0, 0]
    benign_num = 0
    protocol = df.loc[:, "Protocol"].values
    for k in range(len(label)):
        if label[k] == "Benign":
            benign_num += 1
            if protocol[k] == 6:
                TCP_NUM[0] += 1
            elif protocol[k] == 17:
                UDP_NUM[0] += 1
            else:
                pass
        else:
            if protocol[k] == 6:
                TCP_NUM[1] += 1
            elif protocol[k] == 17:
                UDP_NUM[1] += 1
            else:
                pass
    print(len(protocol))
    # 计算总体的TCP和UDP数据包占比
    Total_TCP = (float)(TCP_NUM[0] + TCP_NUM[1]) / len(protocol)
    Total_UDP = (float)(UDP_NUM[0] + UDP_NUM[1]) / len(protocol)
    # 计算正常和异常情况下TCP和UDP数据包的占比
    if benign_num!=0:
        B_ratio_TCP = (float)(TCP_NUM[0]) / benign_num
        B_ratio_UDP = (float)(UDP_NUM[0]) / benign_num
    else:
        B_ratio_TCP=B_ratio_UDP=0
    if (len(protocol) - benign_num)!=0:
        N_ratio_TCP = (float)(TCP_NUM[1]) / (len(protocol) - benign_num)
        N_ratio_UDP = (float)(UDP_NUM[1]) / (len(protocol) - benign_num)
    else:
        N_ratio_TCP=N_ratio_UDP=0
    # 移除"Benign"标签
    if "Benign" in label_type:
        label_type.remove("Benign")

    return len(label_type),Total_TCP, B_ratio_TCP, N_ratio_TCP, Total_UDP, B_ratio_UDP, N_ratio_UDP


def mkdir(path):
    # os.path.exists 函数判断文件夹是否存在
    path = path + '\index'
    folder = os.path.exists(path)
    # 判断是否存在文件夹如果不存在则创建为文件夹
    if not folder:
        # os.makedirs 传入一个path路径，生成一个递归的文件夹；如果文件夹存在，就会报错,因此创建文件夹之前，需要使用os.path.exists(path)函数判断文件夹是否存在；
        os.mkdir(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
        print('文件夹创建成功')
    else:
        print('文件夹已经存在')
    return path

if __name__ == "__main__":

    # 在path路径下创建一个新文件夹
    #Root_path = r'F:\code'
    Root_path=os.getcwd()
    # 创建一个新的文件夹用于保存结果
    path = mkdir(Root_path)
    # 创建一个空的DataFrame用于保存分析结果
    df = pd.DataFrame(
            {
                "Timestamp":[],
                "Total_TCP": [],
                "B_ratio_TCP": [],
                "N_ratio_TCP": [],
                "Total_UDP": [],
                "B_ratio_UDP": [],
                "N_ratio_UDP": [],
                "T_ratio_Small":[],
                "T_ratio_Big":[],
                "B_ratio_Small": [],
                "B_ratio_Big": [],
                "N_ratio_Small": [],
                "N_ratio_Big": [],
                "T_DownUp_Ratio":[],
                "B_DownUp_Ratio":[],
                "N_DownUp_Ratio":[],
                "T_PktLen":[],
                "B_PktLen": [],
                "N_PktLen": [],
                "T_Flowdur":[],
                "B_Flowdur": [],
                "N_Flowdur": [],
                "N_freq": [],
                "N_type":[]
            }
        )
    df.to_excel(path + "\index.xlsx")  # 地址按具体的进行改写
    # 获取指定目录下的所有CSV文件
    input_dir=Root_path+"\\Slice"
    csv_list = glob.glob(input_dir+r'\*.csv')
    #print(csv_list)
    # 遍历每个CSV文件进行分析
    count = 2
    for each_csv in csv_list:
        print("****************************处理文件"+each_csv+"****************************")
        file_path = each_csv
        # 读取CSV文件中的数据并进行分析
        df = pd.read_csv(file_path)
        timestamp = df.loc[:,"Timestamp"][0]
        print(timestamp)
        Fre_mal2,PktTu1, PktTu2, PktTu3, PktTu4, PktTu5, PktTu6 = PktTu(file_path)
        T_ratio_Small,T_ratio_Big,PktAvg1, PktAvg2, PktAvg3, PktAvg4 = PktAvg(file_path)
        T_DownUp_Ratio, B_DownUp_Ratio, N_DownUp_Ratio=Down_Up_Ratio(file_path)
        T_PktLen,PktLenStd1, PktLenStd2 = PktLenStd(file_path)
        T_Flowdur,FlowDur1, FlowDur2 = FlowDur(file_path)
        Fre_mal1 = Fre_mal(file_path)
        num_list = [timestamp, PktTu1, PktTu2, PktTu3, PktTu4, PktTu5, PktTu6, T_ratio_Small,T_ratio_Big,PktAvg1, PktAvg2, PktAvg3, PktAvg4, T_DownUp_Ratio, B_DownUp_Ratio, N_DownUp_Ratio,T_PktLen,PktLenStd1, PktLenStd2, T_Flowdur,FlowDur1, FlowDur2, Fre_mal1,Fre_mal2]
        # 打开之前创建的Excel文件，将分析结果写入其中
        bg = op.load_workbook(path + "\index.xlsx")  # 应先将excel文件放入到工作目录下
        sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
        for i in range(1, len(num_list) + 1):
            sheet.cell(count, i+1, num_list[i - 1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第2行第2列
            bg.save(path + "\index.xlsx")  # 对文件进行保存
        count += 1
    print("done!")


    # print("正常情况下小于等于32字节的数据包占比为：", PktAvg1)
    # print("异常情况下100字节的数据包占比为：", PktAvg2)
    # print("正常情况下等于32字节的数据包占比为：", PktAvg3)
    # print("异常情况下100字节的数据包占比为：", PktAvg4)
    # print("正常情况下流的标准差长度平均值为：", PktLenStd1)
    # print("异常情况下的标准差长度平均值为：", PktLenStd2)
    # print("正常情况下流量持续时间的平均值为：", FlowDur1)
    # print("异常情况下流量持续时间的平均值为：", FlowDur2)
    # print("恶意攻击的频率为：", Fre_mal1)
    # print("TCP协议包总占比：", PktTu1)
    # print("正常情况TCP包占比：", PktTu2)
    # print("异常情况TCP包占比：", PktTu3)
    # print("UDP协议包总占比：", PktTu4)
    # print("正常情况UDP包占比：", PktTu5)
    # print("异常情况UDP包占比：", PktTu6)
